import os
import io
import jwt
import uuid
import json
import docx
import openpyxl
import tempfile
import tarfile
from datetime import datetime, timedelta
import logging

import tornado.web
import tornado.httputil

from handlers.eva.base import BaseHandler

__author__ = "Fedor Metelkin"
__copyright__ = "Copyright 2020, ISG Neuro"
__credits__ = []
__license__ = ""
__version__ = "0.0.1"
__maintainer__ = "Andrey Starchenkov"
# TODO тут почта мэйнтейнера. Так что ты уже реши, кто крайний за модуль.
__email__ = "fmetelkin@isgneuro.com"
__status__ = "Develop"

# TODO Каталог должен быть вынесен в конфиг, а не захардкожен.
# TODO PEP8 иначе казнить, нельзя помиловать.
# TODO Это тоже из PEP8, но все равно отдельно вынесу. Отступы только в пробелах и соответсвенно уровням вложенности. Ужас.
# TODO Все зависимости внешние в requirements.txt
# TODO Ты зачем все файлы перезаписал в ветке?

class PaperLoadHandler(BaseHandler):   # метод отвечающяя за загрузку файл в папку
    def initialize(self, **kwargs):  # инициализируем переменные которые придут при вызове этого метода
        super().initialize(kwargs['db_conn_pool'])
        self.static_conf = kwargs['static_conf']
        self.logger = logging.getLogger('osr')

    async def prepare(self): #  метод без которого не будет работать класс, в нем мы проверяем куку и разрешение на дальнейшую работу
        client_token = self.get_cookie('eva_token')
        if client_token:
            self.token = client_token
            try:
                token_data = self.decode_token(client_token)
                user_id = token_data['user_id']
                self.permissions = self.db.get_permissions_data(user_id=user_id,
                                                                names_only=True)
            except (jwt.ExpiredSignatureError, jwt.DecodeError):
                pass
            else:
                self.current_user = user_id

        if not self.current_user:
            raise tornado.web.HTTPError(401, "unauthorized")

    async def post(self): # метод который положит файл в папку
        body = self.request.body # получаем данные с фронта
        args = {}
        files = {}
        tornado.httputil.parse_body_arguments(self.request.headers['Content-Type'], body, args, files) # парсим и получаем данные в нужном нам виде

        # TODO нельзя склеивать пути руками, используй либу
        reports_path = self.static_conf['static_path'] + 'reports' # путь куда будем сохранять файл
        # TODO добавить проверку на то, что файл там вообще есть, иначе вылетишь с ошибкой. Сообщить пользователю, если файла нет.
        _file = files['file'][0] # тут из всех переданных данных забираем собственно файл

        saving_full_path = os.path.join(reports_path, _file['filename']) # тут к пути еще добовляем имя файла
        # if not os.path.exists(saving_full_path): # тут не до конца понимаю как работает но в общем проверят есть ли уже такой файл и если нет
        # TODO добавить try/catch для отлавливания ошибок записи и возвращения нормального фидбэка пользователю
        with open(saving_full_path, 'wb') as f: # то открывает его (и создает видимо)
            f.write(_file['body'])  # и записывает в него данные с фронта
        self.write({'status': 'success'}) # передаем успешное выполнение запроса

# TODO а тут где проверка авторизации? Добавить.
class PapersHandler(BaseHandler):  # метод возвращающий все файлы в папке на фронт
    def initialize(self, **kwargs):   # инициализируем переменные которые придут при вызове этого метода
        super().initialize(kwargs['db_conn_pool'])
        self.static_conf = kwargs['static_conf']
        self.logger = logging.getLogger('osr')

    async def get(self): # метод который вернет все файлы в папке
      reports_path = self.static_conf['static_path'] + 'reports'  # путь до нужной папки
      files = []  # переменная в которой будет храниться список файлов
      for file in os.listdir(reports_path): # забираем список всех файлов и каталогов из нужной папки
        if os.path.isfile(os.path.join(reports_path, file)): # проверяем если это файл, а не каталог
          files.append(file) # то заносим этот файл в подготовленный массив
      # listOfFiles = [f for f in os.listdir(reports_path) if os.path.isfile(f)] ну почти.. прием хороший, но дебажить его не возможно
      # print(files)
      
      self.write({'files':files,'status': 'success'}) # возвращаем список файлов и сообщение что успешно все прошло 


class PaperHandler(BaseHandler): # метод который изменит файл с фротна и вернет ссылку на новый
    def initialize(self, **kwargs): # инициализируем переменные которые придут при вызове этого метода
        super().initialize(kwargs['db_conn_pool'])
        self.static_conf = kwargs['static_conf']
        # TODO ты нигде дальше не используешь self.mem_conf. Зачем сохранять его в тело класса?
        self.mem_conf = kwargs['mem_conf']
        self.data_path = self.mem_conf['path']
        self.logger = logging.getLogger('osr')
        self.static_dir_name = 'storage'

    async def prepare(self):   # инициализируем переменные которые придут при вызове этого метода
        client_token = self.get_cookie('eva_token')
        if client_token:
            self.token = client_token
            try:
                token_data = self.decode_token(client_token)
                user_id = token_data['user_id']
                self.permissions = self.db.get_permissions_data(user_id=user_id,
                                                                names_only=True)
            except (jwt.ExpiredSignatureError, jwt.DecodeError):
                pass
            else:
                self.current_user = user_id

        if not self.current_user:
            raise tornado.web.HTTPError(401, "unauthorized")
        

    async def post(self): # метод который изменит файл на основе данных с фронта и отдаст ссылку на новый
        body = self.request.body # получаем данные с фронта
        args = {}
        # TODO ты нигде не используешь переменную data, а потом перезаписываешь ее. Отсюда можно убрать.
        data = {}
        tornado.httputil.parse_body_arguments(self.request.headers['Content-Type'], body, args, {}) # парсим и получаем данные в нужном нам виде
        file_name = args['file'][0].decode('utf-8')  # получаем имя нужного файла после раскадировки
        # TODO никогда не склеивай пути руками. Черевато последствиями.
        reports_path = self.static_conf['static_path'] + 'reports' # путь до папки откуда брать файлы

        try:  # попытаемся получить готовые данные с фронта
          data = args['data'][0].decode('utf-8') # если получить удалось
          data = {"status": "success", "data": json.loads(data)} # то подготавливаем дату в нужный нам вид и записываем туды данные
        # TODO нужно конретизировать ошибку, т.к. тут могут быть проблемы парсинга data. То есть она пришла, но ты не смог прогрузить JSON.
        # В этом случае не верно считать, что она не пришла.
        except: # если такого ключа нет, значит нам эти данные нужно поулчить самим
          cid = args['cid'][0].decode('utf-8')  # получаем cid запроса
          data = self.get_data(cid)  # вызываем метод для получения данных
          # TODO Датасет может быть очень большим. Ты сожрешь всю память и ОСь прибьет процесс. Ты должен ввести ограничение хотя бы на количество строк.
          for i, json_data in enumerate(data['data']): # так же нам надо перевести строки json в dist  поэтому пробегаемся по всем данным
            # TODO В этот момент ты увеличил в 2.5 раза объем занимаемый датасетом. Зачем?
            data['data'][i] = json.loads(json_data)  # и распаршиваем json данные в dict

          
        # TODO Когда тебя спросят, почему не пересчитали запрос сами, говори, что его результаты могли измениться
        #  за прошедшее время, и пользователь должен на них снова посмотреть.
        if data['status'] == 'failed':
          self.write({'description':'cache is cleared and search is gone','status': 'failed'})
        else:
          data = data['data']

          # TODO Снова зря инициализировал переменную, т.к. дальше ее перезаписываешь.
          file_res = []  # результат в виде ссылки на измененный файл
          full_path =  os.path.join(reports_path,  file_name)  # полный путь уже с именем нужного файла
          about_file =  file_name.split('.') # получаем массив в котором первое значение имя файла, а второе его разрешение

          # TODO Ни одной попытки поймать ощибку ниже нет. Так нельзя. Там проблем может быть тьма.
          if about_file[1] == 'xlsx': # если файл с разрешением xlsx 
            file_res = self.work_xlsx(full_path,data,about_file[0]) # то вызываем метод который обработает xlsx
          # TODO а если файл будет не docx? Обязательно проверку и тут и вывод ошибки, если файл другой.
          else:
            file_res =self.work_docx(full_path,data,about_file[0]) # то вызываем метод который обработает docx

          self.write({'file':file_res,'status': 'success'}) # вернем ссылку на новый обработанный файл и сообщение что все успешно прошло

    def work_docx(self,path,data,name_file): # метод для работы с xlsx файлами
      # TODO По-хорошему, надо вынести в отдельный модуль. Преобразование файлов не должно быть частью хэндлера.
      result = ''
      files= []
      reports_path = self.static_conf['static_path'] + 'reports/changed'  # задаем правлиьный путь для измененных файлов

      for i, part_data in enumerate(data):

        

        doc = docx.Document(path)

        for paragraph in doc.paragraphs:  
          for key in part_data.keys(): # пробегаемся по словарю данных с фронта
            # TODO Зачем тебе проверка на наличие? Сразу заменяй. Если параграф не содержит совпдаений, то останется неизменным.
            # Если совпадения будут, то будет произведена замена.
            if  paragraph.text.find('$'+key+'$') != -1: # а затем проверяем есть ли в этой ячейке ключ словаря
              paragraph.text = paragraph.text.replace('$'+key+'$', part_data[key])

        if (len(data) > 1):   # если несколько строк данных
          # TODO Дай нормальные имена переменным. file_name и name_file. За что ты так с нами?
          filename = f"{name_file}_{datetime.strftime(datetime.now()+ timedelta(seconds=i), '%Y%m%d%H%M%S')}.docx" # то создаем несоклько файлов но каждому следующему увеличиваем время на секунду
        else: # если строка всего одна
          filename = f"{name_file}_{datetime.strftime(datetime.now(), '%Y%m%d%H%M%S')}.docx" # то просто задаем ей имя исходя из времени создания

        files.append(filename) # уже полный путь с названием файла
        doc.save(os.path.join(reports_path, filename))  # сохраняем измененный файл в папку

      if len(files) > 1: #  если у нас несколько файлов

        # TODO никогда не склеивай пути руками.
        result = 'reports/changed/'+self.to_archive(name_file,files,reports_path) # задаем путь до архива, вызвав метод для создания архивов

      else:  # если файл только один
        result = f"reports/changed/{name_file}_{datetime.strftime(datetime.now(), '%Y%m%d%H%M%S')}.docx" # просто указываем путь до архива

      return result # возвращаем ссылку на измененный файл 

    def work_xlsx(self,path,data,name_file): # метод для работы с xlsx файлами
      # TODO По-хорошему, надо вынести в отдельный модуль Преобразование файлов не должно быть частью хэндлера.

      files = []
      reports_path = self.static_conf['static_path'] + 'reports/changed'  # задаем правлиьный путь для измененных файлов
      result = ''

     
      # TODO Очень странно перебираешь ты данные) Почитай другую статью)
      for i, part_data in enumerate(data):

      

        wb = openpyxl.load_workbook(path) # открываем файл
        sheet = wb.active  # выбираем активный лист
        # TODO Зачем? Ты его нигде не используешь дальше.
        sheet_name = wb.get_sheet_names()[0] # здесь запоминаме имя этого активного листа
  
        for rownum in range(sheet.max_row): # пробегаемся по всем строкам 
          for columnnum in range(sheet.max_column): #  и в каждой строке по всем столбцам
            # TODO А что в итоге с первой строкой будет? И что будет когда ты прибавишь 1 к последней строке?
            cell = sheet.cell(rownum + 1, columnnum + 1).value #  запоминаем занчение в текущей ячейки
            # TODO По факту ты повторяешься здесь. Вынеси  в отдельный метод.
            for key in part_data.keys(): # пробегаемся по словарю данных с фронта
              if cell is not None and type(cell) is str: # првоеряем не пустая ли ячейка и что ячейка строка 
                if  cell.find('$'+key+'$') != -1: # а затем проверяем есть ли в этой ячейке ключ словаря
                  cell = cell.replace('$'+key+'$', part_data[key])  # то заменяем значение ячейке на значение из данных
                  sheet.cell(rownum + 1, columnnum + 1).value = cell # Записываем измененую ячейку в файл
        
        if (len(data) > 1): # если несколько строк данных
          filename = f"{name_file}_{datetime.strftime(datetime.now()+ timedelta(seconds=i), '%Y%m%d%H%M%S')}.xlsx" # то создаем несоклько файлов но каждому следующему увеличиваем время на секунду
        else: # если строка всего одна
          filename = f"{name_file}_{datetime.strftime(datetime.now(), '%Y%m%d%H%M%S')}.xlsx"  # то просто задаем ей имя исходя из времени создания

        files.append(filename) # уже полный путь с названием файла
        wb.save(os.path.join(reports_path, filename)) # сохраняем измененный файл в папку

      if len(files) > 1: #  если у нас несколько файлов

        result = 'reports/changed/'+self.to_archive(name_file,files,reports_path) # задаем путь до архива, вызвав метод для создания архивов

      else:  # если файл только один
        result = f"reports/changed/{name_file}_{datetime.strftime(datetime.now(), '%Y%m%d%H%M%S')}.xlsx" # просто указываем путь до архива

  

      return result # возвращаем ссылку на измененный файл 


    def to_archive(self,name_file,files,reports_path):
      
      with tempfile.TemporaryDirectory() as directory: # создаем временную папку ## ну ты и лентяй

        archive_name = f"{name_file}_{datetime.strftime(datetime.now(), '%Y%m%d%H%M%S')}_archive.tar"
        archive_path = os.path.join(directory, archive_name)  # задаем путь до архива во временной папке
        # TODO Уверен, что на винде откроется архив этот?
        archive = tarfile.open(archive_path, mode='x:gz')  # открываем архив

        for name in files:  # пробегаемся по всем файлам
          os.rename(os.path.join(reports_path, name), os.path.join(directory, name)) # перемещаем созданные файлы во временную папку

        # TODO По-моему, ты имя файла неверное указываешь
        archive.add(directory, name_file+'-changed') # добовляем их в архив

        archive.close() # закрываем архив
        os.rename(os.path.join(directory, archive_name), os.path.join(reports_path, archive_name)) # переносим архив в папку с изменнными файлами

      return archive_name

    def get_data(self, cid): # метод для поулчения данных запроса
      result = {} # здесь будет результат выполнения метода

      try:   # попробуем
        path_to_cache_dir = os.path.join(self.data_path, f'search_{cid}.cache/data') #  получить путь до файла, и если удалось
        file_names = [file_name for file_name in os.listdir(path_to_cache_dir) if file_name[-5:] == '.json'] # то берем все файлы с разрешением json
        length = len(file_names)  # смотрим сколько в итоге файлов
        # TODO Ты усложняешь перебор по списку. Используй оптимальный способ.
        for i in range(length): # пробегаемся по массиву файлов
            file_name = file_names[i] # достаем искомый файл
            with open(os.path.join(path_to_cache_dir, file_name)) as fr: # открываем его для прочтения
                # TODO Размер файла может превышать объем нашей памяти. Нельзя читать его целиком. Нужно ввести ограничение на объем или читать построчно.
                body = fr.read().strip().split('\n')  # считываем содержимое файлов, избавляясь от пустых мест, и сразу разбивая его в массив по строкам
        result = {"status": "success","data": body}  #отдаем успешный статус и наши данные
      except: # если не получилось достучаться до файла
        result = {"status": "failed"}  # то скорее всего время жизни кэша истекло и его больше нет, возвращаем статус failed

      # TODO Тут просто докапываюсь. Так обычно так не делают. В питоне чаще отдают кортеж (True, body) или (False, )
      # А на выходе проверяют первое булевое значение. Работает в десятки раз быстрее твоего способа и жрет меньше памяти.
      return result
      



